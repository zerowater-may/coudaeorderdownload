import datetime
import os
import re
import threading
import webbrowser
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import time

# import pan/das as pd
# from pandas import DataFrame as df
from config import ConfigFile
import openpyxl
from gui.db.conndb import SqliteDb
from gui.uis.libs.customlibrary import Custom_Library
from gui.uis.windows.main_window.defpyqt import DefPyQt
from .googlespread import GoogleSpreadSheet
from qt_core import *
from gspread import Cell, Worksheet

class SyncSpreadSheetData(QThread):
    update_msg = Signal(str)

    def __init__(self,sheetinfo,onlyinvoice=False, parent=None):
        super().__init__()
        self.parent = parent
        self.onlyinvoice = onlyinvoice
        self.sheetinfo = sheetinfo

        self.custom = Custom_Library()
        self.config = ConfigFile()
        self.sheet = GoogleSpreadSheet()
        self.sql = SqliteDb()
        self.default = DefPyQt()
        
        self.application_path = self.config.read()['system']['dir']

    def nowtime(self):
        now = datetime.datetime.now()
        return now.strftime("%Y/%m/%d %H:%M:%S")
    def cleanText(self,readData):
        
        #텍스트에 포함되어 있는 특수 문자 제거
        text = re.sub('[-=+,#/\?:^$.@*\"※~&%ㆍ!』\\‘|\(\)\[\]\<\>`\'…》]','', readData)
        text = re.sub(r'\s+','', text)
        text = text.replace(' ','').replace('  ','').replace('   ','').strip().rstrip().lstrip()
        text = text.upper()
        text = text[:18]
        return text
    
    def run(self):
        '''시트 정보와 실제 데이터를 조회해서 데이터를 수정함'''
        if self.onlyinvoice: self.update_msg.emit('운송장을 엑셀 폼으로 생성합니다. 시트 3번인 상태 , 택배사 , 운송장번호가 기입 된 행을 엑셀로 변환합니다.') 
        else: self.update_msg.emit('스프레드시트 동기화를 시작합니다.')

        S = {
        'A': '1', 'B': '2','C': '3','D': '4','E': '5','F': '6','G': '7','H': '8','I': '9','J': '10','K': '11','L': '12','M': '13','N': '14','O': '15','P': '16','Q': '17','R': '18','S': '19','T': '20','U': '21','V': '22','W': '23','X': '24','Y': '25','Z': '26',
        'AA': '27', 'AB': '28','AC': '29','AD': '30','AE': '31','AF': '32','AG': '33','AH': '34','AI': '35','AJ': '36','AK': '37','AL': '38','AM': '39','AN': '40','AO': '41','AP': '42','AQ': '43','AR': '44','AS': '45','AT': '46','AU': '47','AV': '48','AW': '49','AX': '50','AY': '51','AZ': '52',}
        
        self.orderdata  = self.sql.selectSqlData('ShoppingMallOrderList')['result']
        # self.orderdata = [(636, 'A00328835', '네이버쇼핑', '소프트 매트 컴플리트 컨실러', 40000, '2320955553', '2320955553', '배송준비중', '유용상', '10109', '경기도 안양시 만안구 장내로 36 (안양동) 해든 팰리스 C동 201호 ( 안양동 ) ', 'https://order.pay.naver.com/orderStatus/2022112113676011', 'CJ대한통운', '562061683595', '2022.11.21 13:35', '(4045-****-****-****) -', '', 'C:\\Users\\user\\Dropbox\\zerowater\\billy\\DATA\\COOKIE\\네이버쇼핑네이버ybbang0202.pkl', '2022-11-23 01:28:55')]
        self.ordNos = [o[5] for o in self.orderdata]
        # for i in self.orderdata:
        #     input(i)
        # info['sheetname'] = 'TITANS STORE'
        result = []
        for idx,sheet in enumerate(self.sheetinfo):
            self.update_msg.emit(f'현재진행중 -> {sheet[2]} | {sheet[3]}')
            print(idx,sheet)
# self.onlyinvoice
            # 실제 데이터
            sheetname,sheeturl,No,name,wheretobuy,price,courier,invoicenum,ordNo,adressnum,buytime,idnum = sheet[2],sheet[3],int(S[sheet[4]])-1,int(S[sheet[5]])-1,int(S[sheet[6]])-1,int(S[sheet[7]])-1,int(S[sheet[8]])-1,int(S[sheet[9]])-1,int(S[sheet[10]])-1,int(S[sheet[11]])-1,int(S[sheet[12]])-1,int(S[sheet[13]])-1
            # sheetname = 'TITANS STORE'

            # sheeturl = 'https://docs.google.com/spreadsheets/d/1PfD8X3ZTASV22QXnkSIUlau7tYK1dq5nVNZOJMWKbAg/edit#gid=1537881676'
            # 시트 정보 연결하기
            self.worksheet,istrue = self.sheet.getWorksheet(sheeturl,sheetname)
            # print('self.worksheet',self.worksheet,istrue)
            if istrue == False: 
                self.update_msg.emit(f'현재진행중 -> 연동 실패했습니다.\n{self.worksheet}')
                continue
                # return False # 만약 못가져오면 FAlse

            self.update_msg.emit(f'현재진행중 -> 연동 성공했습니다.')


            # 시트 별 데이터 
            sheetdata = self.sheet.getAllValues() # 시트에 모든 데이터를 가져온다.
            # print('sheetdata',sheetdata)
            if sheetdata == False: 
                self.update_msg.emit(f'[{sheet[2]}]의 실제 데이터를 추출 실패했습니다.')
                continue
                # return False # 만약 못가져오면 FAlse

            self.update_msg.emit(f'[{sheet[2]}]의 실제 데이터를 추출 성공했습니다.')

            # 실제로 넣어주는 부분
            for i,v in enumerate(sheetdata):        
                i+=1
                if self.onlyinvoice: # 운송장 업데이트 버튼 전용
                    if v[No] == '': v[No] = 0
                    if v[No] == 'No': continue
                    if v[No] == 3 or v[No] == '3':  # 상태가 2인 것들 동기화 시켜버리기
                        print(v[idnum],v[courier],v[invoicenum])
                        # input(v[No])

                        if v[idnum] != '' and v[courier] != '' and v[invoicenum] != '':
                            # makedata = {'주문고유코드':str(v[idnum]),'송장번호':str(v[invoicenum]),'택배사':str(v[courier])}
                            # print(makedata)
                            makedata = [str(v[idnum]),str(v[invoicenum]),str(v[courier])]
                            result.append(makedata)
                    continue
                
                if v[No] == '': v[No] = 0
                if v[No] == 'No': continue

                # if v[No] == 9 or v[No] == '9':  # 상태가 2인 것들 동기화 시켜버리기
                # if v[No] == 5 or v[No] == '5':  # 상태가 2인 것들 동기화 시켜버리기
                if v[No] == 2 or v[No] == '2':  # 상태가 2인 것들 동기화 시켜버리기
                    # .lstrip().rstrip()
                    cells = []
                    no = 2
                    # print('2번 주문건')
                    # 모든 주문번호
                    print('=>>',v[No],v[name],v[wheretobuy],v[price],v[courier],v[invoicenum],v[ordNo],v[adressnum])
                    # input(v[ordNo])
                    v[ordNo] = str(v[ordNo])
                    # v[ordNo] = '3865381328'
                    if v[ordNo].lstrip().rstrip() == '': 
                        msg = f'현재시트에 주문번호가 미입력\n[{self.nowtime()}]'
                        updatemsg = f'현재시트에 주문번호가 미입력\n[{self.nowtime()}]'
                        no = 9
                        cells.append(Cell(row=i, col=1, value=msg)) # 메세지 
                        cells.append(Cell(row=i, col=No+1, value=no))
                        self.sheet.updateCells(cells)
                        # self.update_msg.emit(updatemsg)

                    # print('@@',v[ordNo].lstrip().rstrip(),d[5].lstrip().rstrip())
                    # input(self.ordNos)
                    # input(v[ordNo])
                    if str(v[ordNo]).lstrip().rstrip() not in self.ordNos:
                        self.update_msg.emit(f'{sheet[2]} - [{i}]행 주문번호 존재하지않습니다.')
                        msg = f'수집/매칭된 주문번호가 없습니다.\n[{self.nowtime()}]'
                        updatemsg = f'수집/매칭된 주문번호가 없습니다.\n[{self.nowtime()}]'
                        no = 9
                        cells.append(Cell(row=i, col=1, value=msg)) # 메세지 
                        cells.append(Cell(row=i, col=No+1, value=no))
                        self.sheet.updateCells(cells)
                        continue

                    for d in self.orderdata:
                        # 만약 주문번호가 같으면
                        error = False
                        if str(v[ordNo]).lstrip().rstrip() == str(d[5]).lstrip().rstrip(): 
                            updatemsg = ''
                            print('주문번호가 같음')
                            msg = f'{d[7]}\n[{self.nowtime()}]'
                            # 우편번호가 틀릴 시
                            print(d[9] , v[adressnum],d[2])
                         
                            # input('우편번호 체크')
                            print(v[wheretobuy])
                            if '티몬' in d[2]:
                                print('티몬이라 우편번호 체크안함')
                            else:
                                if len(v[adressnum]) == 4: v[adressnum] = '0'+v[adressnum] # 4자리 일경우.
                                
                                if '옥션' in d[2] or '지마켓' in d[2]:
                                    print('우편번호 : 옥션이나 지마켓임',d[2])
                                    if d[9][:3] != v[adressnum][:3]:
                                        error = True
                                        msg = f'우편번호가 틀립니다.\n사이트:{d[9]}\n입력값:{v[adressnum]}\n[{self.nowtime()}]'
                                        updatemsg = f'{sheet[2]} - [{i}]행 우편번호가 틀립니다. | 사이트 : {d[9]} -> 입력값: {v[adressnum]}'
                                        no = 9

                                elif d[9] != v[adressnum]:
                                    error = True
                                    msg = f'우편번호가 틀립니다.\n사이트:{d[9]}\n입력값:{v[adressnum]}\n[{self.nowtime()}]'
                                    updatemsg = f'{sheet[2]} - [{i}]행 우편번호가 틀립니다. | 사이트 : {d[9]} -> 입력값: {v[adressnum]}'
                                    no = 9

                            print(d[8] , v[name])
                            # input('구매자 이름이 다를 시')
                            # 구매자 이름이 다를 시
                            if '옥션' in d[2] or '지마켓' in d[2]:
                                print('이름 : 옥션이나 지마켓임',d[2])
                                if self.cleanText(d[8])[0] !=self.cleanText(v[name])[0]:
                                    error = True
                                    msg = f'수취인명이 틀립니다.\n사이트:{d[8]}\n입력값:{v[name]}\n[{self.nowtime()}]'
                                    updatemsg = f'{sheet[2]} - [{i}]행 수취인명이 틀립니다. | 사이트:{d[8]} -> 입력값:{v[name]}'
                                    no = 9

                            elif self.cleanText(d[8]) != self.cleanText(v[name]):
                                error = True
                                msg = f'수취인명이 틀립니다.\n사이트:{d[8]}\n입력값:{v[name]}\n[{self.nowtime()}]'
                                updatemsg = f'{sheet[2]} - [{i}]행 수취인명이 틀립니다. | 사이트:{d[8]} -> 입력값:{v[name]}'
                                no = 9

                            print(v[courier] ,'|', v[invoicenum] ,'|', d[12],'|',d[13])
                            # input('구매자 이름이 다를 시')
                            # 만약 운송장 번호 , 택배사가 나왔을 경우 
                            if v[courier] == '' and v[invoicenum] == '' and d[12] != '' and d[13] != '' and error == False:
                                no = 3
                                updatemsg = f'{sheet[2]} - [{i}]행 운송장 업데이트 | {v[courier]} / {v[invoicenum]}'
                                cells.append(Cell(row=i, col=courier+1, value=d[12])) # No
                                cells.append(Cell(row=i, col=invoicenum+1, value=d[13])) # No
                                                                                      
                            # 최종 넣기
                            cells.append(Cell(row=i, col=No+1, value=no)) # 넘버링 
                            cells.append(Cell(row=i, col=1, value=msg)) # 메세지 
                            cells.append(Cell(row=i, col=2, value=d[11])) # 다이렉트 링크
                            cells.append(Cell(row=i, col=buytime+1, value=d[14])) # 구매시간

                            cells.append(Cell(row=i, col=price+1, value=d[4])) # 가격
                            cells.append(Cell(row=i, col=wheretobuy+1, value=d[17].split('\\')[-1].replace('.pkl',''))) # 

                            self.sheet.updateCells(cells)
                            if updatemsg != '': self.update_msg.emit(updatemsg)

            self.update_msg.emit(f'[{sheet[2]}] 동기화가 완료되었습니다.')

        if self.onlyinvoice: # 운송장 업데이트 버튼 전용
            excelform = 'shopmine'
            print(result)
            # pd.options.display.float_format = '{:.5f}'.format
            # input(result)
            # ff = df(result)
            # ff.astype(str)
            # df.applymap(str)
            # df.astype(str)
            # df.astype('string')
            # df.applymap(lambda x: x[0] if type(x) is list else None)

            if len(result) == 0: self.update_msg.emit('운송장 입력된 주문건이 0건 입니다.'); return
            
            # 엑셀 폴더 경로
            excelpath = os.path.join(self.application_path,'INVOICEUPLOAD')
            formattedDate = str(self.custom.now().strftime("%Y%m%d_%H%M%S"))

            try: 
                self.custom.createFolder(excelpath)
                webbrowser.open(excelpath)
            except Exception:
                self.default.showMessageBoxs('관리자 권한 오류','관리자 권한이 없는 폴더입니다.\n프로그램 폴더위치를 바탕화면으로 옮겨주세요.')
            wb = openpyxl.Workbook()
            # wb.create_sheet("TEST")
            sheet = wb['Sheet'] # 시트 선택
            header = ['주문고유코드', '송장번호', '택배사']
            for i, value in enumerate(header):
                sheet.cell(row=1, column=i+1, value=value)

            for db in result:
                sheet.append(db)

            wb.save(os.path.join(excelpath,f'운송장업로드_{excelform}_{formattedDate}_{len(result)}건')+'.csv')
            
            # ff.to_csv(
            #     os.path.join(excelpath,
            #     f'운송장업로드_{excelform}_{formattedDate}_{len(result)}건')+'.csv',encoding='utf-8-sig',index=False)
            
            self.update_msg.emit(f'생성완료 | 운송장 입력된 주문건이 {len(result)}건 입니다.')
                            




    